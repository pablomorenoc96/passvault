"""Modelo de datos de la bóveda: entradas, CRUD, importación y exportación."""
from __future__ import annotations

import csv
import os
import secrets
import unicodedata
import uuid
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path

from . import crypto
# Se re-exportan para que quien use la bóveda no tenga que importar crypto.
from .crypto import BovedaCorrupta, ContrasenaIncorrecta, ErrorBoveda

__all__ = ["Boveda", "Entrada", "ErrorBoveda", "ContrasenaIncorrecta",
           "BovedaCorrupta", "buscar_excel_para_migrar"]

MAX_HISTORIAL = 10
CAMPOS_EXPORT = ["Sitio", "Usuario", "Contraseña", "URL", "Categoría", "Notas",
                 "Creado", "Modificado"]


def _ahora() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _texto(valor) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip()
    return "" if texto.lower() == "nan" else texto


@dataclass
class Entrada:
    sitio: str = ""
    usuario: str = ""
    contrasena: str = ""
    url: str = ""
    categoria: str = ""
    notas: str = ""
    favorito: bool = False
    id: str = field(default_factory=lambda: uuid.uuid4().hex)
    creado: str = field(default_factory=_ahora)
    modificado: str = field(default_factory=_ahora)
    historial: list = field(default_factory=list)

    @classmethod
    def desde_dict(cls, datos: dict) -> "Entrada":
        conocidos = {f for f in cls.__dataclass_fields__}
        limpio = {k: v for k, v in datos.items() if k in conocidos}
        entrada = cls(**limpio)
        entrada.favorito = bool(entrada.favorito)
        if not isinstance(entrada.historial, list):
            entrada.historial = []
        return entrada

    def como_dict(self) -> dict:
        return asdict(self)

    def coincide(self, texto: str) -> bool:
        texto = texto.lower()
        return any(texto in campo.lower() for campo in
                   (self.sitio, self.usuario, self.url, self.categoria, self.notas))


class Boveda:
    """Bóveda cifrada en memoria, respaldada por un único archivo en disco."""

    def __init__(self, ruta: Path, contrasena: str, datos: dict | None = None,
                 kdf: dict | None = None):
        self.ruta = Path(ruta)
        self._contrasena = contrasena
        self._kdf = kdf or crypto.parametros_kdf()
        datos = datos or {}
        self.entradas: list[Entrada] = [
            Entrada.desde_dict(item) for item in datos.get("entradas", [])
        ]
        self.creada = datos.get("creada", _ahora())

    # ---------------------------------------------------------------- ciclo de vida
    @classmethod
    def crear(cls, ruta: Path, contrasena: str) -> "Boveda":
        ruta = Path(ruta)
        if ruta.exists():
            raise ErrorBoveda(f"Ya existe una bóveda en {ruta}")
        boveda = cls(ruta, contrasena)
        boveda.guardar()
        return boveda

    @classmethod
    def abrir(cls, ruta: Path, contrasena: str) -> "Boveda":
        ruta = Path(ruta)
        respaldo = Path(str(ruta) + ".bak")

        try:
            contenido = ruta.read_bytes()
            datos, kdf = crypto.descifrar(contenido, contrasena)
            return cls(ruta, contrasena, datos, kdf)
        except (crypto.BovedaCorrupta, OSError) as exc_principal:
            # Si el archivo principal se dañó (por ej. corte de energía), intentar rescatar desde .bak
            if respaldo.exists():
                try:
                    contenido_bak = respaldo.read_bytes()
                    datos, kdf = crypto.descifrar(contenido_bak, contrasena)
                    boveda_recuperada = cls(ruta, contrasena, datos, kdf)
                    boveda_recuperada.guardar()  # Restaura el archivo principal automáticamente
                    return boveda_recuperada
                except Exception:
                    pass  # Si el respaldo tampoco abre, reportar el fallo del principal
            raise ErrorBoveda(f"No se pudo leer la bóveda: {exc_principal}") from exc_principal

    def guardar(self) -> None:
        datos = {
            "creada": self.creada,
            "modificada": _ahora(),
            "entradas": [e.como_dict() for e in self.entradas],
        }
        contenido = crypto.cifrar(datos, self._contrasena, self._kdf)
        try:
            crypto.escribir_atomico(self.ruta, contenido)
        except OSError as exc:
            raise ErrorBoveda(f"No se pudo guardar la bóveda: {exc}") from exc

    def cambiar_maestra(self, nueva: str) -> None:
        anterior = self._contrasena
        self._contrasena = nueva
        self._kdf = crypto.parametros_kdf()  # sal y parámetros nuevos
        try:
            self.guardar()
        except Exception:
            self._contrasena = anterior
            raise

    def verificar_maestra(self, contrasena: str) -> bool:
        if self._contrasena is None:
            return False
        c1 = unicodedata.normalize("NFC", contrasena).encode("utf-8")
        c2 = unicodedata.normalize("NFC", self._contrasena).encode("utf-8")
        return secrets.compare_digest(c1, c2)

    # ---------------------------------------------------------------------- CRUD
    def obtener(self, id_entrada: str) -> Entrada | None:
        return next((e for e in self.entradas if e.id == id_entrada), None)

    def agregar(self, entrada: Entrada) -> Entrada:
        entrada.creado = entrada.modificado = _ahora()
        self.entradas.append(entrada)
        self.guardar()
        return entrada

    def actualizar(self, id_entrada: str, **campos) -> Entrada:
        entrada = self.obtener(id_entrada)
        if entrada is None:
            raise ErrorBoveda("La entrada ya no existe.")

        nueva_pass = campos.get("contrasena")
        if nueva_pass is not None and nueva_pass != entrada.contrasena and entrada.contrasena:
            entrada.historial.append({"contrasena": entrada.contrasena,
                                      "fecha": entrada.modificado})
            del entrada.historial[:-MAX_HISTORIAL]

        for clave, valor in campos.items():
            if hasattr(entrada, clave) and clave not in ("id", "creado", "historial"):
                setattr(entrada, clave, valor)
        entrada.modificado = _ahora()
        self.guardar()
        return entrada

    def eliminar(self, ids: list[str]) -> int:
        objetivo = set(ids)
        antes = len(self.entradas)
        self.entradas = [e for e in self.entradas if e.id not in objetivo]
        borradas = antes - len(self.entradas)
        if borradas:
            self.guardar()
        return borradas

    def duplicar(self, id_entrada: str) -> Entrada | None:
        original = self.obtener(id_entrada)
        if original is None:
            return None
        copia = Entrada.desde_dict(original.como_dict())
        copia.id = uuid.uuid4().hex
        copia.sitio = f"{original.sitio} (copia)"
        copia.historial = []
        return self.agregar(copia)

    # ------------------------------------------------------------------ consultas
    def filtrar(self, texto: str = "", categoria: str | None = None,
                solo_favoritos: bool = False) -> list[Entrada]:
        resultado = self.entradas
        if solo_favoritos:
            resultado = [e for e in resultado if e.favorito]
        if categoria:
            resultado = [e for e in resultado if (e.categoria or "Sin categoría") == categoria]
        if texto:
            resultado = [e for e in resultado if e.coincide(texto)]
        return resultado

    def categorias(self) -> list[str]:
        vistos = {e.categoria.strip() for e in self.entradas if e.categoria.strip()}
        return sorted(vistos, key=str.lower)

    def contrasenas_repetidas(self) -> dict[str, list[Entrada]]:
        por_clave: dict[str, list[Entrada]] = {}
        for entrada in self.entradas:
            if entrada.contrasena:
                por_clave.setdefault(entrada.contrasena, []).append(entrada)
        return {k: v for k, v in por_clave.items() if len(v) > 1}

    # --------------------------------------------------------------- importación
    def _agregar_lote(self, filas: list[dict], omitir_duplicados: bool = True) -> tuple[int, int]:
        """Agrega varias entradas de una sola vez. Devuelve (agregadas, omitidas)."""
        existentes = {(e.sitio.lower(), e.usuario.lower(), e.contrasena)
                      for e in self.entradas}
        agregadas = omitidas = 0

        for fila in filas:
            entrada = Entrada(
                sitio=_texto(fila.get("sitio")),
                usuario=_texto(fila.get("usuario")),
                contrasena=str(fila.get("contrasena") or ""),
                url=_texto(fila.get("url")),
                categoria=_texto(fila.get("categoria")),
                notas=str(fila.get("notas") or ""),
            )
            if not entrada.sitio and not entrada.usuario:
                omitidas += 1
                continue
            clave = (entrada.sitio.lower(), entrada.usuario.lower(), entrada.contrasena)
            if omitir_duplicados and clave in existentes:
                omitidas += 1
                continue
            existentes.add(clave)
            self.entradas.append(entrada)
            agregadas += 1

        if agregadas:
            self.guardar()
        return agregadas, omitidas

    def importar_texto(self, texto: str, separador: str = ",",
                       omitir_duplicados: bool = True) -> tuple[int, int]:
        """Importa líneas con formato ``Sitio<sep>Usuario<sep>Contraseña[<sep>URL...]``."""
        filas = []
        errores = 0
        for linea in texto.splitlines():
            linea = linea.strip()
            if not linea:
                continue
            partes = [p.strip() for p in linea.split(separador)]
            if len(partes) < 2:
                errores += 1
                continue
            partes += [""] * (5 - len(partes))
            filas.append({"sitio": partes[0], "usuario": partes[1],
                          "contrasena": partes[2], "url": partes[3],
                          "categoria": partes[4]})
        agregadas, omitidas = self._agregar_lote(filas, omitir_duplicados)
        return agregadas, omitidas + errores

    def importar_excel(self, ruta, omitir_duplicados: bool = True) -> tuple[int, int]:
        from openpyxl import load_workbook

        libro = load_workbook(ruta, read_only=True, data_only=True)
        hoja = libro.active
        filas_iter = hoja.iter_rows(values_only=True)
        try:
            encabezado = next(filas_iter)
        except StopIteration:
            libro.close()
            return 0, 0

        indices = _mapear_columnas([_texto(c) for c in encabezado])
        filas = []
        if not indices:
            # Sin encabezado reconocible: se asumen las 3 primeras columnas.
            indices = {"sitio": 0, "usuario": 1, "contrasena": 2}
            filas.append(_fila_a_dict(encabezado, indices))

        for cruda in filas_iter:
            filas.append(_fila_a_dict(cruda, indices))
        libro.close()
        return self._agregar_lote(filas, omitir_duplicados)

    def importar_csv(self, ruta, omitir_duplicados: bool = True) -> tuple[int, int]:
        with open(ruta, "r", encoding="utf-8-sig", newline="") as fh:
            muestra = fh.read(4096)
            fh.seek(0)
            try:
                dialecto = csv.Sniffer().sniff(muestra, delimiters=",;\t")
                dialecto.quotechar = '"'
            except csv.Error:
                dialecto = csv.excel
            lector = csv.reader(fh, dialecto)
            todas = [list(f) for f in lector]

        if not todas:
            return 0, 0

        indices = _mapear_columnas([_texto(c) for c in todas[0]])
        cuerpo = todas[1:]
        if not indices:
            indices = {"sitio": 0, "usuario": 1, "contrasena": 2}
            cuerpo = todas
        filas = [_fila_a_dict(f, indices) for f in cuerpo]
        return self._agregar_lote(filas, omitir_duplicados)

    # --------------------------------------------------------------- exportación
    def exportar_excel(self, ruta) -> int:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        libro = Workbook()
        hoja = libro.active
        hoja.title = "Contraseñas"
        hoja.append(CAMPOS_EXPORT)

        relleno = PatternFill("solid", fgColor="1F3A5F")
        for celda in hoja[1]:
            celda.font = Font(bold=True, color="FFFFFF")
            celda.fill = relleno
            celda.alignment = Alignment(horizontal="center")

        for e in sorted(self.entradas, key=lambda x: x.sitio.lower()):
            hoja.append([_sanitizar_formula(e.sitio),
                         _sanitizar_formula(e.usuario),
                         _sanitizar_formula(e.contrasena),
                         _sanitizar_formula(e.url),
                         _sanitizar_formula(e.categoria),
                         _sanitizar_formula(e.notas),
                         e.creado, e.modificado])

        for columna, ancho in zip("ABCDEFGH", (28, 32, 26, 30, 18, 40, 20, 20)):
            hoja.column_dimensions[columna].width = ancho
        hoja.freeze_panes = "A2"

        libro.save(ruta)
        return len(self.entradas)

    def exportar_csv(self, ruta) -> int:
        with open(ruta, "w", encoding="utf-8-sig", newline="") as fh:
            escritor = csv.writer(fh)
            escritor.writerow(CAMPOS_EXPORT)
            for e in sorted(self.entradas, key=lambda x: x.sitio.lower()):
                escritor.writerow([_sanitizar_formula(e.sitio),
                                   _sanitizar_formula(e.usuario),
                                   _sanitizar_formula(e.contrasena),
                                   _sanitizar_formula(e.url),
                                   _sanitizar_formula(e.categoria),
                                   _sanitizar_formula(e.notas),
                                   e.creado, e.modificado])
        return len(self.entradas)


# ------------------------------------------------------------------- utilidades
ALIAS_COLUMNAS = {
    "sitio": ("sitio", "site", "sitio web", "nombre", "name", "title", "titulo",
              "servicio", "cuenta", "aplicacion", "app"),
    "usuario": ("usuario", "user", "username", "correo", "email", "e-mail",
                "login", "user name"),
    "contrasena": ("contrasena", "contraseña", "password", "pass", "clave",
                   "pwd", "secret"),
    "url": ("url", "web", "link", "enlace", "direccion", "website"),
    "categoria": ("categoria", "categoría", "grupo", "carpeta", "folder",
                  "group", "type", "tipo"),
    "notas": ("notas", "nota", "note", "notes", "comentario", "comments",
              "observaciones"),
}


def _normalizar(texto: str) -> str:
    tabla = str.maketrans("áéíóúÁÉÍÓÚñÑ", "aeiouAEIOUnN")
    return texto.translate(tabla).strip().lower()


def _mapear_columnas(encabezado: list[str]) -> dict[str, int]:
    indices: dict[str, int] = {}
    for posicion, titulo in enumerate(encabezado):
        limpio = _normalizar(titulo)
        for campo, alias in ALIAS_COLUMNAS.items():
            if campo in indices:
                continue
            if limpio in (_normalizar(a) for a in alias):
                indices[campo] = posicion
                break
    # Se exige al menos sitio o usuario para considerarlo un encabezado real.
    return indices if ("sitio" in indices or "usuario" in indices) else {}


def _sanitizar_formula(valor: str) -> str:
    """Previene Formula Injection / CSV Injection (CWE-1236) en hojas de cálculo.
    Si una celda inicia con =, +, -, @, \\t o \\r, se neutraliza con un apóstrofo (')
    para que Excel o Calc lo traten como texto literal seguro."""
    if not valor:
        return ""
    val_str = str(valor)
    if val_str and val_str[0] in ("=", "+", "-", "@", "\t", "\r"):
        return "'" + val_str
    return val_str


def _desanitizar_formula(valor: str) -> str:
    """Restaura el valor original si fue protegido previamente con apóstrofo."""
    if valor and valor.startswith("'") and len(valor) > 1 and valor[1] in ("=", "+", "-", "@", "\t", "\r"):
        return valor[1:]
    return valor


def _fila_a_dict(fila, indices: dict[str, int]) -> dict:
    resultado = {}
    for campo, posicion in indices.items():
        try:
            val = fila[posicion]
            resultado[campo] = _desanitizar_formula(_texto(val))
        except (IndexError, TypeError):
            resultado[campo] = ""
    return resultado


def buscar_excel_para_migrar() -> list[Path]:
    """Excels en texto plano junto al programa, candidatos a importarse."""
    from .config import directorio_app

    carpeta = directorio_app()
    candidatos = []
    for nombre in ("mis_contrasenas.xlsx", "contrasenas_organizadas.xlsx"):
        ruta = carpeta / nombre
        if ruta.exists() and os.path.getsize(ruta) > 0:
            candidatos.append(ruta)
    return candidatos
